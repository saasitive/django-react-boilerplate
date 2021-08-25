from .models import TSites, TSiteconfig, TEventdata, TEnergydata, TEdits, TPowercurves

from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.core import serializers
from django.forms.models import model_to_dict
from django.views.decorators.csrf import csrf_protect
from wopr.utils import *
from datetime import timedelta, datetime
import dateutil.parser
from django.core.serializers.json import DjangoJSONEncoder
from django.core.exceptions import ObjectDoesNotExist
import json
from django.db.models import Sum, Count, Q
from django.core.serializers.json import DjangoJSONEncoder
import copy
import os
import time
import pytz
from django.views.generic import TemplateView
from django.db.models import Min, Max, Value
from django.db import connection
from django.utils import timezone
from django.conf import settings as djangoSettings

from django.contrib.sessions.models import Session
from openpyxl import *
from operator import itemgetter
import pprint
from dateutil.parser import parse
import getpass


# from .forms import siteID_startTime_endTime, CommentForm
from .forms import *
# Create your views here.
# 1. Create index and details method here
# 2. Interact with model here
# 3. Pull our posts in
# 4. Load templates and views from here


# ------- Functions used to display the html with a certain dictionary ------- #

# Sets the timezone of the user by storing the timezone name in django_timezone which system will determine offset with
def set_timezone(request):

    if request.method == 'POST':
        if(request.POST.get('timezone')):

            tzname = request.POST.get('timezone')

            request.session['django_timezone'] = tzname
            request.session['local'] = tzname

            timezone.activate(pytz.timezone(tzname))
            print('TimezoneMiddleware activated timezone to', tzname, 'after POST.')

        return JsonResponse({"message": 'The timezone was successfully set to: '}, safe=False)
        
    else:
        return JsonResponse({"message": 'The request has to be a POST to set the timezone.'}, safe=False)

# Main page to select a site
def selectSite(request):
    
    # Check if the user is visiting the page
    if request.method == 'GET':

        # Resets the selected site everytime you click on select a site
        for key in request.session.keys():
            if key != 'django_timezone' and key != '_csrftoken':
                request.session[key] = ''
            
        form = SiteSelectionForm()
        return render(request, 'wopr/selectSite.html', {'form': form, 'title': 'Select a Site'})
        
        
    # Check if the request is a post
    if request.method == 'POST':
        form = SiteSelectionForm(request.POST)
        if form.is_valid():

            siteid = form.cleaned_data['site_id']
            sites_queryset = TSites.objects.filter(siteid__icontains=siteid).order_by('siteid')

            form = SiteSelectionForm()

            return HttpResponseRedirect('/wopr/dashboard/' + str(siteid))
            
        return render(request, 'wopr/selectSite.html', {'form': form,  'title': 'Select a Site', 'message': 'Please make sure to enter a valid site id.'})

# Get the site with the siteID 'site'
def getTSites(request):

    # Check whether the data parameters are valid
    if not(isNum(request.GET['siteid'])) or not(request.GET['siteid']):
        return JsonResponse({'status':'0', 'message': 'Please enter a valid site ID.'})
        
    # Check if the request is ajax and a type: "GET"
    if(request.is_ajax and request.GET):
        # Get the data parameters
        data = request.GET
        # Search the database where siteid = data['siteid']
        try:
            sites = TSites.objects.get(siteid=data['siteid']) # use Model.objects.get() if there is one object in the database or None

            # Convert the object to a dictionary with a format:
            # (example)
            #   {
            #       'id': "1",
            #       'name': "Bob"
            #   }
            #
            dictionary = model_to_dict(sites)

            # Add status and message to the dictionary
            dictionary['status'] = "1"
            dictionary['message'] = "Success"

            # Use sessions to store the selected site id
            request.session['siteid'] = data['siteid']
            # Set the sessions expiry to 10 minutes
            #request.session.set_expiry(600)

            # Send back the JSON containing the data of sites to the AJAX call
            # with a Content-Type header set to application/json 
            
            # The safe boolean parameter defaults to True. If it’s set to False, any object can be passed for serialization (otherwise only dict instances are allowed). 
            # If safe is True and a non-dict object is passed as the first argument, a TypeError will be raised.
            return JsonResponse(dictionary, safe=False)

        # Catch exception if the site does not exist
        except TSites.DoesNotExist:
            
           return JsonResponse({'status':'0','message':'The site number does not exist in the database'})
        
    return JsonResponse({'status':'0', 'message': 'Request must be Ajax of type "GET"'})

# Get the events with the siteID 'site'
def getEvents(request, site):

    t_events = TEventdata.objects.filter(siteid=site, id=1).order_by('ts_start')

    # Serialize to JSON format
    dataFromTemplate = serializers.serialize("json", t_events, default = dateConverter)

    return render(request, 'wopr/eventsData.html', {'json': dataFromTemplate})
    
# Get the energy data with the siteID 'site'
def getEnergy(request, site):

    t_energy = TEnergydata.objects.filter(siteid=site, id=1).order_by('ts')

    # Serialize to JSON format
    dataFromTemplate = serializers.serialize("json", t_energy, default = dateConverter)

    return render(request, 'wopr/energyData.html', {'json': dataFromTemplate})

# Get the site configuration for the siteID 'site'
def getSiteConfig(request, site):

    t_siteconfig = TSiteconfig.objects.filter(siteid=site).order_by('id')    

    # Serialize to JSON format
    dataFromTemplate = serializers.serialize("json", t_siteconfig, default = dateConverter)

    # Send back the JSON containing the data of t_sites to the AJAX call
    # with a Content-Type header set to application/json 
    return render(request, 'wopr/siteconfig.html', {'json': dataFromTemplate})    

# Get the edits table data for the siteID 'site'
def getEdits(request, site):

    t_edits = TEdits.objects.filter(siteid=site)    

    # Serialize to JSON format
    dataFromTemplate = serializers.serialize("json", t_edits, default = dateConverter)

    # Send back the JSON containing the data of t_sites to the AJAX call
    # with a Content-Type header set to application/json 
    return render(request, 'wopr/editsTable.html', {'json': dataFromTemplate})

# Get the changes array saved in session
def getChangesArray(request):
    if(request.is_ajax and request.method == 'GET'):
        return JsonResponse({"changesArray": request.session['changesArray'], "status": 1}, safe=False)
    else:
        return JsonResponse({"message": 'Please make sure your ajax request is a valid GET.', "status": 1}, safe=False)

# Get the site selection form for the dashboard page
def dashboardSelectSite(request):

    # Check if the user is visiting the page
    if request.method == 'GET':

        # Resets the selected site everytime you click on select a site
        for key in request.session.keys():
            if key != 'django_timezone' and key != '_csrftoken':
                request.session[key] = ''
            
        siteSelectionForm = SiteSelectionForm()
        return render(request, 'wopr/dashboard.html', {'siteSelectionForm': siteSelectionForm, 'title': 'Edits Allocation'})
        
        
    # Check if the request is a post
    if request.method == 'POST':
        siteSelectionForm = SiteSelectionForm(request.POST)
        if siteSelectionForm.is_valid():

            siteid = siteSelectionForm.cleaned_data['site_id']
            sites_queryset = TSites.objects.filter(siteid__icontains=siteid).order_by('siteid')

            siteSelectionForm = SiteSelectionForm()

            return HttpResponseRedirect('/wopr/dashboard/' + str(siteid))
            
        return render(request, 'wopr/dashboard.html', {'siteSelectionForm': siteSelectionForm,  'title': 'Edits Allocation'})

# Get the edits allocation turbine data for the siteID 'site'
def dashboard(request, site):
    start = time.time()

    # Get the colors corresponding to state and system from utils.py
    stateColorScheme = getStateColors()
    systemColorScheme = getSystemColors()
    system_description_list = [system for system in TSystems.objects.all().values('systemid', 'system', 'description', 'defined_state', 'definedstateid')]
    for system_type in systemColorScheme:
        system_type['descriptions'] = []
        for description in system_description_list:
            if(system_type['system'] == description['system']):
                system_type['descriptions'].append(description)

    # if it is get (initial visit) then make a blank form and render it.
    if request.method == 'GET':

        request.session['siteid'] = site # store the siteid in session

        # User wants to select another time range
        if('reselect' in request.GET):

            # Resets the selected site everytime you click on select a site
            for key in request.session.keys():
                if key != 'django_timezone' and key != '_csrftoken' and key != 'siteid':
                    request.session[key] = ''

            form = TurbineSelectionForm(siteid=site)
            # When there are no turbines in the database for that site
            if(form.CHOICES[0][0] == -1 and form.CHOICES[0][1] == 'There are no turbines available for this site.'):
                siteform = SiteSelectionForm()
                return render(request, 'wopr/selectSite.html', {'message': 'There were no turbines in the database for that site.', 'title': 'Select a Site', 'form': siteform})
            
            return render(request, 'wopr/dashboard.html', {'form': form})
        
        # when the user has filled out the filterdashboard form
        elif request.session.get('ts_start_global') and request.session.get('ts_end_global') and request.session.get('turbine_range'): 

            form = TurbineSelectionForm(siteid=site)

            id_from, id_till = request.session['turbine_range'].split('-')

            currentDir = os.getcwd()
            with open(currentDir + '/backend/server/wopr/templates/wopr/locations/siteLocations.json') as file:
                site_locations = json.load(file)

            with open(currentDir + '/backend/server/wopr/templates/wopr/locations/turbineLocations.json') as file:
                turbine_locations = json.load(file)

            filtered_site = {}

             # Create the filterdashboard form
            filterDashboardForm = FilterDashboardForm(siteid=site, id_from=id_from, id_till=id_till)

            end = time.time()
            print('Call to dashboard took ', (end - start) , " seconds and {} queries.".format(len(connection.queries)))

            # Return the dictionary with the list of turbine information            
            return render(request, 'wopr/dashboard.html', {'form': form, 'ts_start_global': request.session.get('ts_start_global'), 'ts_end_global': request.session.get('ts_end_global'), 'turbine_range': request.session.get('turbine_range'),
                            'filterDashboardForm': filterDashboardForm, 'site_locations': json.dumps(site_locations,cls=DjangoJSONEncoder), 'turbine_locations': json.dumps(turbine_locations, cls=DjangoJSONEncoder), 
                            'stateColorScheme': json.dumps(stateColorScheme, cls=DjangoJSONEncoder), 'systemColorScheme': json.dumps(systemColorScheme, cls=DjangoJSONEncoder)})
            
        else:
            form = TurbineSelectionForm(siteid=site)

            # Resets the selected site everytime you click on select a site
            for key in request.session.keys():
                if key != 'django_timezone' and key != '_csrftoken' and key != 'siteid':
                    request.session[key] = ''

            # When there are no turbines in the database for that site
            if(form.CHOICES[0][0] == -1 and form.CHOICES[0][1] == 'There are no turbines available for this site.'):
                siteform = SiteSelectionForm()
                return render(request, 'wopr/selectSite.html', {'message': 'There were no turbines in the database for that site.', 'title': 'Select a Site', 'form': siteform})
            
            return render(request, 'wopr/dashboard.html', {'form': form})

    # if it is a post request (from same page) unpack the form and do things with it.
    if request.method == 'POST':
        form = TurbineSelectionForm(request.POST, siteid=site)
        filterDashboardForm = FilterDashboardForm(request.POST, siteid=site, id_from=request.POST.get('id_from'), id_till=request.POST.get('id_till'))
        
        if form.is_valid():
            # if it was valid unpack the posted values
            turbine_range = str(form.cleaned_data['view_turbines_from']) + '-' + str(form.cleaned_data['view_turbines_till'])

            ts_start_global = form.cleaned_data['start_time']
            ts_end_global = form.cleaned_data['end_time']

            print('Form received: ', ts_start_global, ts_end_global, turbine_range)

            # Map
            currentDir = os.getcwd()
            with open(currentDir + '/backend/server/wopr/templates/wopr/locations/siteLocations.json') as file:
                site_locations = json.load(file)

            with open(currentDir + '/backend/server/wopr/templates/wopr/locations/turbineLocations.json') as file:
                turbine_locations = json.load(file)

            filtered_site = {}

            # Create the filterdashboard form
            filterDashboardForm = FilterDashboardForm(siteid=site, id_from=form.cleaned_data['view_turbines_from'], id_till=form.cleaned_data['view_turbines_till'])

            end = time.time()
            print('Call to dashboard took ', (end - start) , " seconds and {} queries.".format(len(connection.queries)))

            # Return the dictionary with the list of turbine information            
            return render(request, 'wopr/dashboard.html', {'form': form, 'ts_start_global': ts_start_global.__str__, 'ts_end_global': ts_end_global.__str__, 'turbine_range': turbine_range, 'filterDashboardForm': filterDashboardForm,
                            'site_locations': json.dumps(site_locations,cls=DjangoJSONEncoder), 'turbine_locations': json.dumps(turbine_locations, cls=DjangoJSONEncoder), 
                            'stateColorScheme': json.dumps(stateColorScheme, cls=DjangoJSONEncoder), 'systemColorScheme': json.dumps(systemColorScheme, cls=DjangoJSONEncoder)})

        elif filterDashboardForm.is_valid():
            return render(request, 'wopr/dashboard.html', {'form': form, 'filterDashboardForm': filterDashboardForm})
        # (To raise validation error from form, do not make blank form - get rid of else statement)        
        else: # if the form was not valid
            # form = TurbineSelectionForm(siteid=site)
            return render(request, 'wopr/dashboard.html', {'form': form})

def ajaxTenMinuteData(request):
    if(request.method == "GET"):
        start = time.time()

        if(request.GET.get('tz')):
            request.session['django_timezone'] = request.GET.get('tz')

        fmt = '%Y-%m-%d %H:%M:%'

        ts_start_UTC = dateutil.parser.parse(request.GET.get('ts_start_from')).astimezone(pytz.utc)
        ts_until_UTC = dateutil.parser.parse(request.GET.get('ts_until')).astimezone(pytz.utc)

        turbine_range_from, turbine_range_till = request.GET.get('turbine_range').split("-")
        
        turbine_range_from = int(turbine_range_from)
        turbine_range_till = int(turbine_range_till)

        if ts_start_UTC > ts_until_UTC:
            return JsonResponse({'status': 0, 'message': "Ensure the date ranges are valid."}, safe=False)

        if turbine_range_from > turbine_range_till:
            return JsonResponse({'status': 0, 'message': "Ensure that the turbine range is valid."}, safe=False)

        site = request.GET.get('siteid')

        site_turbines = TSiteconfig.objects.filter(siteid=site, id__range=[turbine_range_from, turbine_range_till]).order_by('id').values('id', 'kksname')# Queryset
        rawDataList = []

        # Iterate through the QuerySet
        for turbine in site_turbines:

            dictionary = {
                'site': site,
                'turbine': turbine['id'],
                'KKSName': turbine['kksname'], 
                'data': ''
            }
            fetch_eventdata = TSiteconfig.objects.filter(siteid=site, id=turbine['id']).prefetch_related("teventdata_set")
            fetch_energydata = TSiteconfig.objects.filter(siteid=site, id=turbine['id']).prefetch_related("tenergydata_set")

            for each_turbine_event, each_turbine_energy in zip(fetch_eventdata, fetch_energydata):

                # Replace .all() to filter()

                energy_query_set = each_turbine_energy.tenergydata_set.all().order_by("ts").exclude(kw_net__isnull=True).exclude(kw_exp__isnull=True).exclude(nws__isnull=True)
                event_query_set = each_turbine_event.teventdata_set.all().order_by("ts_start")

                
                innerList = []
                current_state = ''
                for each_event in event_query_set:
                    energy_data = [d for d in energy_query_set if d.periodid == each_event.periodid]
                
                    if(len(energy_data) > 0):
                            inner_dictionary = {**model_to_dict(energy_data[0]), **model_to_dict(each_event)}
                            # Convert all the datetime.datetime objects to a string with format "YYYY-MM-DD HH:MM:SS"
                            for key in inner_dictionary:
                                inner_dictionary[key] = dateConverter(inner_dictionary[key])
                            innerList.append(inner_dictionary)

                dictionary['data'] = innerList

            rawDataList.append(dictionary)
        
        if not len(rawDataList):
            return JsonResponse({'status': 0, 'message': "There were no data for the date ranges."}, safe=False)
            
        end = time.time()
        print('It took ', (end - start) , " seconds and {} queries.".format(len(connection.queries)))

        # JSON serializer

        return JsonResponse({'status': 1,'rawDataList': json.dumps(rawDataList,cls=DjangoJSONEncoder), 'ts_start_global': '2012-03-09 00:00:00+00:00', 'ts_end_global': '2012-03-16 00:00:00+00:00'})

    return JsonResponse({'status': 0, 'message': "Your data request was invalid."}, safe=False)


# ------- Functions used to filter the tables to show filtering with different parameters ------- #


# Filter the event table for specific columns of ts_start_from and ts_until
def filterEvent(request):

    dataParams = ''

    # Check if the request is Ajax and is of type GET
    if(request.is_ajax() and request.GET):
        dataParams = request.GET    # set the parameters of the request GET to dataParams for easier naming

        # Case when we dont want to revert back to displaying all
        if dataParams['ts_start_from'] == '' and dataParams['ts_until'] == '':
            t_events = TEventdata.objects.filter(siteid=dataParams['siteid'], id=1)

            # Serialize to JSON format
            dataFromTemplate = serializers.serialize("json", t_events, default = dateConverter)

            return JsonResponse(dataFromTemplate, safe=False)

        aware_ts_start = dateutil.parser.parse(dataParams['ts_start_from']).astimezone(pytz.timezone(request.session.get('django_timezone')))
        aware_ts_end = dateutil.parser.parse(dataParams['ts_until']).astimezone(pytz.timezone(request.session.get('django_timezone')))

        # print(aware_ts_start)

        # Get the ts_start entries of each end
        t_events = TEventdata.objects.filter(siteid=dataParams['siteid'],
            ts_start__range=(aware_ts_start,
                            aware_ts_end), id=1
        ).order_by('ts_start')

        # Serialize to JSON format
        dataFromTemplate = serializers.serialize("json", t_events, default = dateConverter)

        # Send back the JSON containing the data of t_sites to the AJAX call
        # with a Content-Type header set to application/json 
    
        return JsonResponse(dataFromTemplate, safe=False)

# Filter the energy table for specific columns of ts_start_from and ts_until
def filterEnergy(request):

    dataParams = ''

    # Check if the request is Ajax and is of type GET
    if(request.is_ajax() and request.GET):
        dataParams = request.GET    # set the parameters of the request GET to dataParams for easier naming

        # Case when we dont want to revert back to displaying all
        if dataParams['ts_start_from'] == '' and dataParams['ts_until'] == '':
            t_energy = TEnergydata.objects.filter(siteid=dataParams['siteid'], id=1)

            # Serialize to JSON format
            dataFromTemplate = serializers.serialize("json", t_energy, default = dateConverter)

            return JsonResponse(dataFromTemplate, safe=False)

        aware_ts_start = dateutil.parser.parse(dataParams['ts_start_from']).astimezone(pytz.timezone(request.session.get('django_timezone')))
        aware_ts_end = dateutil.parser.parse(dataParams['ts_until']).astimezone(pytz.timezone(request.session.get('django_timezone')))

        # print(aware_ts_start)

        # Fetch the energy data between the two time stamps ts_start_from, and ts_until
        t_energy = TEnergydata.objects.filter(siteid=dataParams['siteid'],ts__range=(
            aware_ts_start, aware_ts_end), id=1)

        # Serialize to JSON format
        dataFromTemplate = serializers.serialize("json", t_energy, default = dateConverter)

        # Send back the JSON containing the data of t_sites to the AJAX call
        # with a Content-Type header set to application/json 
    
        return JsonResponse(dataFromTemplate, safe=False)

# Filter the site config table for specific column of Generator Make
def filterConfig(request):

    dataParams = ''

    # Check if the request is Ajax and is of type GET
    if(request.is_ajax() and request.GET):
        dataParams = request.GET    # set the parameters of the request GET to dataParams for easier naming

        # Case when we dont want to revert back to displaying all
        if dataParams['genMake'] == '':
            t_siteconfig = TSiteconfig.objects.filter(siteid=dataParams['siteid'])

            # Serialize to JSON format
            dataFromTemplate = serializers.serialize("json", t_siteconfig, default = dateConverter)

            return JsonResponse(dataFromTemplate, safe=False)

        # Fetch all the entries from the TSiteConfig table with the Generator make 'genMake'
        t_siteconfig = TSiteconfig.objects.filter(siteid=dataParams['siteid'],generator_make=dataParams['genMake'])

        # Serialize to JSON format
        dataFromTemplate = serializers.serialize("json", t_siteconfig, default = dateConverter)

        # Send back the JSON containing the data of t_sites to the AJAX call
        # with a Content-Type header set to application/json 
    
        return JsonResponse(dataFromTemplate, safe=False)

def editsTableView (request):
    template_name = 'wopr/editsTable2.html'
    edits_table_form = EditsTableForm()
    site_form = SiteSelectionForm()
    if request.method == 'POST':
        if 'submit-site' in request.POST:
            site_form = SiteSelectionForm(request.POST)
            if site_form.is_valid():
                site_id = site_form.cleaned_data['site_id']
                return render(request, template_name, {'site_form': site_form, 'site_selected':site_id, 'edits_table_form':edits_table_form})
        if 'submit-range-user' in request.POST:
            edits_table_form = EditsTableForm(request.POST)
            if edits_table_form.is_valid():
                site_id = edits_table_form.cleaned_data['site_selected']
                start_time = edits_table_form.cleaned_data['start_time']
                end_time = edits_table_form.cleaned_data['end_time']
                user = edits_table_form.cleaned_data['usernames']
                print(site_id, end_time, user)
                t_edits = TEdits.objects.filter(siteid=site_id, ts_edit__gte=start_time, ts_edit__lte=end_time, username=user)
                # Serialize to JSON format
                #data = json.dumps(t_edits)
                data = serializers.serialize('json', t_edits)
                #print(data)
                return render(request, template_name, {'site_selected':site_id, 'edits_table_form':edits_table_form, 'data':t_edits})
    else:
        site_form = SiteSelectionForm()
    return render(request, template_name, {'site_form': site_form, 'edits_table_form':edits_table_form})

# Filter the edits table for specific columns of username
def filterEdits(request):

    dataParams = ''

    # Check if the request is Ajax and is of type GET
    if(request.is_ajax() and request.GET):
        dataParams = request.GET    # set the parameters of the request GET to dataParams for easier naming

        # Case when we dont want to revert back to displaying all
        if dataParams['username'] == '':
            t_edits = TEdits.objects.filter(siteid=dataParams['siteid'])

            # Serialize to JSON format
            dataFromTemplate = serializers.serialize("json", t_edits, default = dateConverter)

            return JsonResponse(dataFromTemplate, safe=False)

        # Fetch all the entries from the TSiteConfig table
        t_edits = TEdits.objects.filter(siteid=dataParams['siteid'],username=dataParams['username'])

        # Serialize to JSON format
        dataFromTemplate = serializers.serialize("json", t_edits, default = dateConverter)

        # Send back the JSON containing the data of t_sites to the AJAX call
        # with a Content-Type header set to application/json 
    
        return JsonResponse(dataFromTemplate, safe=False)

# ------- Functions used to save the changes to the database ------- #

def commitPage(request):

    # Get the colors corresponding to state and system from utils.py
    stateColorScheme = getStateColors()
    systemColorScheme = getSystemColors()
    system_description_list = [system for system in TSystems.objects.all().values('systemid', 'system', 'description', 'defined_state', 'definedstateid')]
    for system_type in systemColorScheme:
        system_type['descriptions'] = []
        for description in system_description_list:
            if(system_type['system'] == description['system']):
                system_type['descriptions'].append(description)

    return render(request, 'wopr/commitPage.html', {'title': 'Commit Page', 'ts_start_global': request.session.get('ts_start_global'), 'ts_end_global': request.session.get('ts_end_global'), 'turbine_range': request.session.get('turbine_range'),
     'stateColorScheme': json.dumps(stateColorScheme, cls=DjangoJSONEncoder), 'systemColorScheme': json.dumps(systemColorScheme, cls=DjangoJSONEncoder)})

@csrf_protect
def postDashboardChanges(request):

    if (request.is_ajax and request.POST):
        dataParams = request.POST
        for key in request.POST.keys():
            print(key)

        request.session['changesArray'] = dataParams['changesArray']

        if not len(request.session['changesArray']):
            return JsonResponse({"message": 'The changes were discarded. Reloading the page.'}, safe=False)
        

        return JsonResponse({"message": 'The changes were successfully saved!\n Your changes can be viewed via the commits difference page.'}, safe=False)

@csrf_protect
def storeSessionData(request):

    if (request.is_ajax and request.POST and request.POST.keys()):

        for key in request.POST.keys():
            request.session[key] = request.POST.get(key)

        return JsonResponse({"message": 'The data was successfully stored in user session.'}, safe=False)

    return JsonResponse({"message": 'Invalid request. Make sure it is ajax of type POST with more than one key.'}, safe=False)

# Commit the changes made from the edits and call the store procedures
def commitDashboardChanges(request):

    if (request.is_ajax):

        if(request.session['changesArray']):
            changesList = json.loads(request.session['changesArray']) 
            #print(changesList)

            stateColorScheme = getStateColors()
            systemColorScheme = getSystemColors()
            system_description_list = [system for system in TSystems.objects.all().values('systemid', 'system', 'description', 'defined_state', 'definedstateid')]
            for system_type in systemColorScheme:
                system_type['descriptions'] = []
                for description in system_description_list:
                    if(system_type['system'] == description['system']):
                        system_type['descriptions'].append(description)

            for change in changesList:

                if change['column'] == 'state': # check if the change was a state change

                    stateInfo = filter(lambda stateDict: stateDict['state'] == change['newValue'], stateColorScheme)
                    for e in stateInfo:
                        ts_start = dateutil.parser.parse(change['ts_start']).astimezone(pytz.utc).strftime('%d-%b-%Y %H:%M')
                        ts_end = dateutil.parser.parse(change['ts_end']).astimezone(pytz.utc).strftime('%d-%b-%Y %H:%M')
                        try:
                            turbine_object = TSiteconfig.objects.get(siteid=request.session.get('siteid'), id=change['turbine'])
                            turbine = getattr(turbine_object, 'turbine')
                            username = getpass.getuser()
                            sql = "exec sp_UpdateStateEdits "+ str(request.session.get('siteid')) + ", '" + turbine + "', '" + ts_start + "', '" + ts_end + "', '" + e['state'] + "', '" + username  + "', ''"
                            print(sql)

                            # with connection.cursor() as cursor:
                                # cursor.execute(sql) # execute the stored procedures
                                
    
                        except TSiteconfig.DoesNotExist:
                            print('Error. Siteid and turbine combination does not exist in TSiteConfig')
                            pass

                elif change['column'] == 'system': # else, a system change

                    systemInfo = filter(lambda systemDict: systemDict['description'] == change['newValue'], system_description_list)
                    for e in systemInfo:
                        ts_start = dateutil.parser.parse(change['ts_start']).astimezone(pytz.utc).strftime('%d-%b-%Y %H:%M')
                        ts_end = dateutil.parser.parse(change['ts_end']).astimezone(pytz.utc).strftime('%d-%b-%Y %H:%M')
                        try:
                            turbine_object = TSiteconfig.objects.get(siteid=request.session.get('siteid'), id=change['turbine'])
                            turbine = getattr(turbine_object, 'turbine')
                            username = getpass.getuser()
                            sql = "exec sp_UpdateSystemEdits "+ str(request.session.get('siteid')) + ", '" + turbine + "', '" + ts_start + "', '" + ts_end + "', '" + e['description'] + "', '" + username  + "', ''"
                            print(sql)

                            # with connection.cursor() as cursor:
                                # cursor.execute(sql) # execute the stored procedures
                                
    
                        except TSiteconfig.DoesNotExist:
                            print('Error. Siteid and turbine combination does not exist in TSiteConfig')
                            pass

            # Then clear the changesArray
            if(request.session['changesArray']):
                del request.session['changesArray']
                
            return JsonResponse({"message": 'The changes were successfully committed!'}, safe=False)

        else:
            return JsonResponse({"message": 'There are no changes to be committed.'}, safe=False)

def discardDashboardChanges(request):
    if (request.is_ajax):

        if(request.session['changesArray']):
            del request.session['changesArray'] 
            return JsonResponse({"message": 'The changes were discarded.'}, safe=False)

    return JsonResponse({"message": 'Invalid request. Make sure it is ajax of type POST with more than one key.'}, safe=False)
    

# ------- Reports views ------- #

def getEditsReport(request):

    '''
    bruteforce method:
    SELECT [editID],[ts_edit],[description], [Turbine], [period_from],[period_to],[ts_EditStart],[ts_EditEnd],[username],[comment]
    FROM [myWOPR].[dbo].[t_edits]
    JOIN [myWOPR].[dbo].[t_sites] ON [myWOPR].[dbo].[t_edits].[siteID] = [myWOPR].[dbo].[t_sites].[SiteID]
    JOIN [myWOPR].[dbo].[t_SiteConfig] ON [myWOPR].[dbo].[t_edits].[id] = [myWOPR].[dbo].[t_SiteConfig].[ID]
    GO
    '''

    # get raw
    # with connection.cursor() as cursor:
    #    cursor.execute(
    '''
        SELECT [editID],[ts_edit],[description], [Turbine], [period_from],[period_to],[ts_EditStart],[ts_EditEnd],[username],[comment]
        FROM [t_edits]
        JOIN [t_sites] ON [t_edits].[siteID] = [t_sites].[SiteID]
        JOIN [t_SiteConfig] ON [t_edits].[id] = [t_SiteConfig].[ID]
        ORDER BY [editID] DESC
    '''
    #    )
    #    unorderedDict = dictfetchall(cursor)
    unorderedDict = {}

    return render(request, 'wopr/reports/editsReport.html', {'quality_report':unorderedDict})

def QualityCheckView(request):
    template_name = 'wopr/reports/editsQualityCheckReport.html'

    # if it is get (initial visit) then make a blank form and render it.
    if request.method == 'GET':
        # sitesDictionary = getUniqueSiteIdAndDescription()
        # print(sitesDictionary)
        form = SiteDateTimeForm()
        return render(request, template_name, {'form':form})

    # if it is a post request (from same page) unpack the form and do things with it.
    if request.method == 'POST':
        form = SiteDateTimeForm(request.POST)
        if form.is_valid():
            # if it was valid unpack the posted values
            site_id = form.cleaned_data['site_id']
            start_time = form.cleaned_data['start_time']
            end_time = form.cleaned_data['end_time']

            # get the tables
            occurrenceData = getOccurrenceTableData(site_id, start_time, end_time)




            args = {'form':form, 'occurrences_report':occurrenceData}
            return render(request, template_name, args)
          
        else:
            # if the form was not valid just blankify the form.
            form = SiteDateTimeForm()
            return render(request, template_name, {'form':form})
    
def top20FaultsView(request):
    template_name = 'wopr/reports/top20FaultsReport.html'
    
    if request.method == 'POST':
        # unpack filled form
        form = SiteTimeRangeForm(request.POST)
        if form.is_valid():
            # selection from form
            site_id = form.cleaned_data['site_id']
            start_time = form.cleaned_data['start_time']
            end_time = form.cleaned_data['end_time']
            # For each eventid, get total duration, convert ms to hours

            top20_Duration_Hr0 = TEventdata.objects.filter(siteid__exact=site_id, ts_start__gte=start_time, ts_end__lte=end_time).values('eventid').distinct()#.annotate(totalDuration = (Sum('duration_ms')/360000)).order_by('-totalDuration')
            top20_Duration_Hr=[]
            for item in top20_Duration_Hr0:
                event = item['eventid']
                d = {}
                d['eventid'] = event
                durationsQuery =TEventdata.objects.filter(siteid__exact=site_id, eventid__exact=event,  ts_start__gte=start_time, ts_end__lte=end_time).values('duration_ms')
                totalDuration = 0
                for item in durationsQuery:
                    totalDuration += item['duration_ms']
                d['totalDuration'] = round(totalDuration/3600000)
                top20_Duration_Hr.append(d)
            # For each eventid, get number of turbines concerned()
            top20_NumTurbines = TEventdata.objects.filter(siteid__exact=site_id, ts_start__gte=start_time, ts_end__lte=end_time).values('eventid').annotate(numTurbines = Count('id', distinct = True))
            #top20 = top20_NumTurbines.union(top20_Duration_Hr)
            durationsJson = json.dumps(list(top20_Duration_Hr), cls=DjangoJSONEncoder)
            turbinesJson = json.dumps(list(top20_NumTurbines), cls=DjangoJSONEncoder)
            if (top20_Duration_Hr and top20_NumTurbines):
                args = {'form':form, 'selection':{'site_id':site_id, 'start_time':start_time, 'end_time':end_time }, 'durationsJson': durationsJson, 'turbinesJson': turbinesJson}
            else: 
                args = {'form':form, 'selection':{'site_id':site_id, 'start_time':start_time, 'end_time':end_time }, 'durationsJson': {}, 'turbinesJson': {} }
            return render(request, template_name, args)
    else:
        # blank form
        form = SiteTimeRangeForm()
    return render(request, template_name, {'form':form})
    
def recurringFaultsView(request):
    #eventid 2 excluded - must check: normal event?
    template_name = 'wopr/reports/recurringFaultReport.html'
    # blank form
    if request.method == 'GET':
        form = SiteTimeRangeForm()
        return render(request, template_name, {'form':form})
    # unpack filled form
    if request.method == 'POST':
        form = SiteTimeRangeForm(request.POST)
        if not form.is_valid(): 
            # if the form was not valid return blank form
            form = SiteTimeRangeForm()
            return render(request, template_name, {'form':form})
        else:
            # selection from form
            site_id = form.cleaned_data['site_id']
            start_time = form.cleaned_data['start_time']
            end_time = form.cleaned_data['end_time']
            # get list of turbines for this site
            setOfTurbines = TEventdata.objects.filter(siteid__exact=site_id).values('id').distinct().order_by('id')
            turbinesIdList = []
            for i in setOfTurbines:
                if not (i['id']==None):
                    turbinesIdList.append(int(i['id']))
            # for each turbine in this site, count distinct events
            turbinesData = {}
            for i in turbinesIdList:
                turbineSet = TEventdata.objects.filter(siteid__exact=site_id, id__exact=i, ts_start__gte=start_time, ts_end__lte=end_time).values('eventid').exclude(eventid=2).distinct().annotate(count=Count('eventid')).order_by('-count') 
                if turbineSet:
                    turbinesData[str(i)]=turbineSet[0]
                    mostFrequentEvent = turbineSet[0]['eventid']
                    firstTimeSet = TEventdata.objects.filter(siteid__exact=site_id, id__exact=i, ts_start__gte=start_time, ts_end__lte=end_time, eventid__exact=mostFrequentEvent, ts_start__isnull=False).aggregate(Min('ts_start'))#.get('ts_start').ealiest('ts_start')#.filter(dates not null)
                    turbinesData[str(i)]['tsMin']=firstTimeSet['ts_start__min']
                    lastTimeSet = TEventdata.objects.filter(siteid__exact=site_id, id__exact=i, ts_start__gte=start_time, ts_end__lte=end_time, eventid__exact=mostFrequentEvent, ts_end__isnull=False).aggregate(Max('ts_end'))
                    turbinesData[str(i)]['tsMax']=lastTimeSet['ts_end__max']
            # get list of events for this site in this time range
            setOfEvents = TEventdata.objects.filter(siteid__exact=site_id, ts_start__gte=start_time, ts_end__lte=end_time).values('eventid').distinct()
            eventsIdList = []
            for i in setOfEvents:
                if not(i['eventid']==2):
                    eventsIdList.append(int(i['eventid']))
            # make a dictionary of dictionaries for the number of occurences of each event for each turbine
            eventsData = {}            
            for i in eventsIdList:
                eventsData[str(i)] = []
                for item in turbinesIdList:
                    eventsData[str(i)].append({'id': item, 'count': ' '})
            # for each event, get number of occurences per turbine
            for i in eventsIdList:
                eventSet = TEventdata.objects.filter(siteid__exact=site_id, eventid__exact=i, ts_start__gte=start_time, ts_end__lte=end_time).values('id').distinct().annotate(count=Count('eventid'))
                for item in eventSet:
                    for dicItem in eventsData[str(i)]:
                        if (dicItem['id'] == item['id']): 
                            dicItem['count'] = item['count']
            # give proper turbine names
            namedTurbineData = {}
            namedturbinesIdList = []
            turbineNames = TSiteconfig.objects.filter(siteid__exact=site_id).values('id','kksname')
            if turbineNames:
                for item in turbineNames:
                    for key in turbinesData:
                        if (int(key) == int(item['id'])):
                            namedTurbineData[item['kksname']] = turbinesData[key]
                    '''
                    for key in eventsData:
                        for listItem in eventsData[key]:
                            if (listItem['id'] == item['id']):
                                listItem['id'] = item['kksname']
                    '''
                    for turbineID in turbinesIdList:
                        if (turbineID == int(item['id'])):
                            namedturbinesIdList.append(item['kksname'])
            # send all data to template
            if namedTurbineData: 
               args = {'form':form, 'selection':{'site_id':site_id, 'start_time':start_time, 'end_time':end_time }, 'turbinesData': namedTurbineData, 'eventsData': eventsData, 'turbinesList':namedturbinesIdList} 
            else: 
                args = {'form':form, 'selection':{'site_id':site_id, 'start_time':start_time, 'end_time':end_time }, 'turbinesData': turbinesData, 'eventsData': eventsData, 'turbinesList':turbinesIdList}
            return render(request, template_name, args)



#need to make a different form to select turbine as well
def powerCurveView(request):
    template_name = 'wopr/reports/powerCurveReport.html'
    # blank form
    if request.method == 'GET': #user hasn't submitted a form (visited the page for the first time)
        form = SiteTurbineTimeForm()
        form2 = CompareTurbinePowerForm()
        return render(request, template_name, {'form':form, 'form2':form2}) #so we return the page with the forms to be filled out
    # unpack filled form
    if request.method == 'POST':
        form = SiteTurbineTimeForm(request.POST)
        if not form.is_valid(): 
            # if the form was not valid return blank form
            form = SiteTurbineTimeForm()
            form2 = CompareTurbinePowerForm()
            return render(request, template_name, {'form':form, 'form2':form2})
        else:
            # selection from the single turbine form was submitted (NOT the compare turbines form, that is submitted to powerCurveCompareView)
            site_id = form.cleaned_data['site_id']
            turbine_id_and_number = form.cleaned_data['turbine_id'] # this turbine id is passed to the form as id-number
            turbine_id = turbine_id_and_number.split("-") # now turbine_id is [id, number]
            start_time = form.cleaned_data['start_time']
            end_time = form.cleaned_data['end_time']
            # For each eventid, get total duration, convert ms to hours
            powerData = TEnergydata.objects.filter(siteid__exact=site_id, ts__gte=start_time, ts__lte=end_time, id__exact = turbine_id[0]).values('ts', 'nws', 'kw_net', 'kw_exp')
            referenceCurve = TPowercurves.objects.filter(siteid__exact=site_id, id__exact = turbine_id[0]).values('nws_bin', 'kw')
            powerJson = json.dumps(list(powerData), cls=DjangoJSONEncoder)
            referenceCurveJson = json.dumps(list(referenceCurve), cls=DjangoJSONEncoder)
            if (powerData):
                args = {'form':form, 'selection':{'site_id':site_id, 'turbine_id':turbine_id[0], 'turbine_name':turbine_id[1], 'start_time':start_time, 'end_time':end_time }, 'powerJson': powerJson, 'referenceCurveJson': referenceCurveJson}
            else: 
                args = {'form':form, 'selection':{'site_id':site_id, 'turbine_id':turbine_id[0], 'turbine_name':turbine_id[1], 'start_time':start_time, 'end_time':end_time }, 'powerJson' : {}, 'referenceCurveJson': {}}
            return render(request, template_name, args)

def filterTurbines(request):
    if (request.is_ajax and request.GET):

        site = request.GET.get('site_id')
        turbines = TSiteconfig.objects.filter(siteid=site).values('id', 'turbine')
        return render(request, 'wopr/formtemplates/turbine_dropdown_options.html', {'turbines': turbines})

    return render(request, 'wopr/formtemplates/turbine_dropdown_options.html', {'turbines': {}})

def powerCurveCompareView(request):
    template_name = 'wopr/reports/powerCurveComparisonReport.html'
    # we wont be getting GET requests here
    if request.method == 'POST':
        form = CompareTurbinePowerForm(request.POST)
        if not form.is_valid(): 
            # if the form was not valid return blank form
            form = CompareTurbinePowerForm()
            return render(request, template_name, {'form':form})
        else:
            # selection from form
            site_id_1 = form.cleaned_data['site_id_1']
            turbine_id_and_number_1 = form.cleaned_data['turbine_id_1'] # this turbine id is passed to the form as id-number
            turbine_id_1 = turbine_id_and_number_1.split("-") # now turbine_id is [id, number]
            start_time_1 = form.cleaned_data['start_time_1']
            end_time_1 = form.cleaned_data['end_time_1']

            site_id_2 = form.cleaned_data['site_id_2']
            turbine_id_and_number_2 = form.cleaned_data['turbine_id_2'] # this turbine id is passed to the form as id-number
            turbine_id_2 = turbine_id_and_number_2.split("-") # now turbine_id is [id, number]
            start_time_2 = form.cleaned_data['start_time_2']
            end_time_2 = form.cleaned_data['end_time_2']
            # For each eventid, get total duration, convert ms to hours
            powerData1 = TEnergydata.objects.filter(siteid__exact=site_id_1, ts__gte=start_time_1, ts__lte=end_time_1, id__exact = turbine_id_1[0]).values('ts', 'nws', 'kw_net', 'kw_exp')
            referenceCurve1 = TPowercurves.objects.filter(siteid__exact=site_id_1, id__exact = turbine_id_1[0]).values('nws_bin', 'kw')
            powerJson1 = json.dumps(list(powerData1), cls=DjangoJSONEncoder)
            referenceCurveJson1 = json.dumps(list(referenceCurve1), cls=DjangoJSONEncoder)

            powerData2 = TEnergydata.objects.filter(siteid__exact=site_id_2, ts__gte=start_time_2, ts__lte=end_time_2, id__exact = turbine_id_2[0]).values('ts', 'nws', 'kw_net', 'kw_exp')
            referenceCurve2 = TPowercurves.objects.filter(siteid__exact=site_id_2, id__exact = turbine_id_2[0]).values('nws_bin', 'kw')
            powerJson2 = json.dumps(list(powerData2), cls=DjangoJSONEncoder)
            referenceCurveJson2 = json.dumps(list(referenceCurve2), cls=DjangoJSONEncoder)

            #pass both sets of data and reference curves

            if (powerData1 and powerData2 and referenceCurve1 and referenceCurve2):
                args = {'form':form, 'selection':{'site_id_1':site_id_1, 'turbine_id_1':turbine_id_1[0], 'turbine_name_1': turbine_id_1[1], 'start_time_1':start_time_1, 'end_time_1':end_time_1, 'site_id_2':site_id_2, 'turbine_id_2':turbine_id_2[0], 'turbine_name_2':turbine_id_2[1], 'start_time_2':start_time_2, 'end_time_2':end_time_2 }, 'powerJson1': powerJson1, 'powerJson2':powerJson2, 'referenceCurveJson1': referenceCurveJson1, 'referenceCurveJson2':referenceCurveJson2}
            else: 
                args = {'form':form, 'selection':{'site_id_1':site_id_1, 'turbine_id_1':turbine_id_1[0], 'turbine_name_1': turbine_id_1[1], 'start_time_1':start_time_1, 'end_time_1':end_time_1, 'site_id_2':site_id_2, 'turbine_id_2':turbine_id_2[0], 'turbine_name_2':turbine_id_2[1], 'start_time_2':start_time_2, 'end_time_2':end_time_2 }, 'powerJson1' : {}, 'powerJson2':{}, 'referenceCurveJson1': {}, 'referenceCurveJson2': {}}
            return render(request, template_name, args)

# ------- IESO file handling ------- #

def iesoView (request):
    template_name = 'wopr/ieso.html'
    if request.method == 'POST':
        if 'file' in request.POST:
            form = UploadIESOFileForm(request.POST, request.FILES)
            if form.is_valid():
                useMidnight = form.cleaned_data['useMidnight']
                iesofile = request.FILES['IESO_file']
                wb = load_workbook(iesofile)
                if "IESO" in wb.sheetnames:
                    wb.active = wb.sheetnames.index("IESO")
                    ws = wb.active
                    if ws['A2'].value:  # if not empty
                        lists = handleUploadedIESOFile(ws, useMidnight)
                        scriptsList = lists[0]
                        errors = lists[1]
                        if errors:
                            args = {'form': form, 'valid': iesofile, 'result': scriptsList, 'errors': errors}
                            return render(request, template_name, args)
                        args = {'form': form, 'valid': iesofile, 'result': scriptsList}
                        return render(request, template_name, args)
                    else:
                        args = {'form': form, 'invalid': iesofile,
                            'result': "seems to contain an 'IESO' sheet with missing data"}
                        return render(request, template_name, args)
                else:
                    args = {'form': form, 'invalid': iesofile,
                        'result': "does not seem to contain an 'IESO' sheet"}
                    return render(request, template_name, args)
        elif 'accepted_scripts' in request.POST:
            form = UploadIESOFileForm()
            s = request.POST.get('accepted_scripts')
            scripts = [x for y in s.split('["') for z in y.split('"]') for x in z.split('", "')]
            # cursor = connection.cursor()

            # The following test works for the data in our database
            # test = "exec sp_UpdateStateSystemWhereCTH 1432, 3, '08-Mar-2012 23:59', '08-Mar-2012 23:59','oDTH','No Data','Flora','dispatched to 53', 1;"
            # cursor.execute(test)


            ''' for item in scripts[1:len(scripts)-1]:
                cmd = "exec sp_UpdateStateSystemWhereCTH " + item
                cursor.execute(cmd)
            cursor.close() '''
            return render(request, template_name, {'form': form})
    else:
        form = UploadIESOFileForm()
    return render(request, template_name, {'form': form})


def handleUploadedIESOFile(ws, useMidnight):
    update_scripts = []
    error_messages = []
    if (ws['A2'].value == None):
        return
    # Find range of Resources table
    max_row = 13
    while (ws['V' + str(max_row)].value):
        max_row += 1
    last_cell = 'Y'+str(max_row-1)
    # Make a dictionary for Resources
    site_table = ws['V14':last_cell]
    resources = {}
    for row in site_table:
        resources[row[0].value] = [row[1].value, row[2].value, row[3].value]
    # Find range of cells
    max_row = 1
    while (ws['A' + str(max_row)].value):
        max_row += 1
    last_cell = 'P'+str(max_row-1)
    # Put data into a list of lists
    data = ws['A1':last_cell]
    rows = []
    for row_number, row in enumerate(data[1:]):
        rows.append([])  # put rows in a list
        for cell in row:
            rows[row_number].append(cell.value)  # each row is a list of cells
    # Sorting rows by 'Resource', then by 'Send Time'
    rows.sort(key=itemgetter(0, 2))
    if not len(rows):
        return update_scripts
    end_row = row_number  # len(rows)
    for row_number, row in enumerate(rows):
        # Clear Start/End column
        row[8] = None
        # If 'MANDATORY', find end date of the dispatch
        if row[5] == 'MANDATORY':
            strDateFrom = row[2]
            if resources:
                intSiteID = resources.get(row[0])[1]
                intgroupID = resources.get(row[0])[2]
            else:
                intSiteID = ''
                intgroupID = ''
            if row[4] > 0:
                state = 'oDTH'
            else:
                state = 'oDTH'  # Cell N2: IF(E2>0,"oDTH","oDTH")
            system = 'Ext - Curtailment'
            comments = 'dispatched to ' + str(round(row[4]))
            row[8] = 'start' # Label as 'start'
            # set start_time_row to compare resource to end time row later
            start_time_row = row_number
            blnAddEndTimeToUnfinishedDispatch = False
            # Find end date for group
            for other_row in rows[row_number+1:]:
                # If Resources are different
                if row[0] != other_row[0]:
                    # end of a group but no release found for dispatch:  Check if they want to use midnight the following day
                    if useMidnight:
                        try:
                            # if strDateFrom is indeed a date: create dtEnd = midnight the day after strDateFrom
                            dtEnd = parse(strDateFrom)
                            dtEnd += timedelta(days=1)
                            dtEnd = dtEnd.replace(hour=0, minute=0, second=0)
                            strDateTo = datetime.strptime(str(dtEnd), '%Y-%m-%d %H:%M:%S').strftime('%d-%b-%Y %H:%M')
                            blnAddEndTimeToUnfinishedDispatch = True
                        except:
                            pass
                    else:
                        errorMsg = "No release found for dispatch at row " + str(row_number+1) +"."
                        error_messages.append(errorMsg)
                        break
                # found the end of a dispatch (or forced to use midnight, as above)
                if (other_row[5] == 'RELEASE' and row[0] == other_row[0]) or blnAddEndTimeToUnfinishedDispatch:
                    if other_row[5] == 'RELEASE' and row[0] == other_row[0]:
                        other_row[8] = 'end'
                        strDateTo = other_row[2]
                    username = getpass.getuser()
                    strSQL = str(intSiteID) + ", " + str(intgroupID) + ", '" + str(
                        strDateFrom) + "', '" + str(strDateTo) + "','" + state + "','" + system + "','" + username + "','" + comments + "', 1;"
                    update_scripts.append(strSQL)
                    strDateTo = ''
                    break
        blnAddEndTimeToUnfinishedDispatch = False
    return ([update_scripts, error_messages])







# def QualityCheckView(TemplateView):
#     template_name = 'wopr/reports/siteStartTimeEndTimeForm.html'

#     def get(self, request):
#         form = SiteDateTimeForm()
#         return render(request, self.template_name, {'form':form})

#     def post(self, request):
#         form = SiteDateTimeForm(request.POST)
#         if form.is_valid():
#             site_id = form.cleaned_data['site_id']
#             start_time = form.cleaned_data['start_time']
#             end_time = form.cleaned_data['end_time']
            
#             args = {'form':form, 'id':site_id, 'st':start_time, 'et':end_time }
#             return render(request, self.template_name, args)
#         else:
#             return render(request, self.template_name, {'form':form})



# function that makes a view for the site, start time, end time, getter form
# def getSiteStartEndTimes(request):
#     template_name = 'wopr/reports/siteStartTimeEndTimeForm.html'
#     if request.method == 'POST':
#         form = siteID_startTime_endTime(request.POST)

#         if form.is_valid():
#             # if the form is valid redirect to new URL:
#             return HttpResponseRedirect('/thanks/')

#         else:
#             # else post a blank form
#             form = siteID_startTime_endTime()

#     def get(self, request):
#         # get the form wen they make a get request:
#         form = siteID_startTime_endTime()
#         return render(request, self.template_name, {'form':form})
#     return render(request, 'wopr/reports/siteStartTimeEndTimeForm.html', {'form':form})

# def conan(request):
#    cursor = connection.cursor()
#    print(pyodbc.version)
#    cursor.execute("exec [dbo].sp_UpdateStateSystemWhereCTH(1528, 3, '28-Jan-2019 19:05', '28-Jan-2019 19:20','oDTH','Ext - Curtailment','Flora','dispatched to 53', 1)")
#    return render(request, 'wopr/404.html')
